import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

wb = openpyxl.Workbook()

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

tables = [
    ("dicom_patient", "Core Hierarchy", "Logical patient entity from DICOM headers (not enterprise MPI)", "Patient → Study → Series → Instance", [
        ("patient_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "(0010,xxxx)", "Y", "N"),
        ("source_system", "VARCHAR", "", "Originating system", "", "N", "N"),
        ("patient_id", "VARCHAR", "NOT NULL", "DICOM Patient ID", "(0010,0020)", "Y", "Y"),
        ("issuer_of_patient_id", "VARCHAR", "", "Issuer of Patient ID", "(0010,0021)", "N", "Y"),
        ("patient_name", "VARCHAR", "", "Patient name", "(0010,0010)", "N", "Y"),
        ("patient_sex", "VARCHAR(16)", "", "M, F, or O", "(0010,0040)", "N", "N"),
        ("patient_birth_date", "DATE", "", "Birth date", "(0010,0030)", "N", "Y"),
        ("patient_age", "VARCHAR(16)", "", "Age at time of study", "(0010,1010)", "N", "N"),
        ("other_patient_ids", "ARRAY", "", "Alternative patient IDs", "(0010,1000)", "N", "Y"),
        ("other_patient_names", "ARRAY", "", "Alternative names", "(0010,1001)", "N", "Y"),
        ("comments", "VARCHAR", "", "Patient comments", "(0010,4000)", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_study", "Core Hierarchy", "Study-level context (order/visit/exam episode)", "FK: patient_key → dicom_patient", [
        ("study_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("patient_key", "INTEGER", "FK → dicom_patient(patient_key)", "Parent patient", "", "Y", "N"),
        ("study_instance_uid", "VARCHAR", "NOT NULL UNIQUE", "Globally unique study identifier", "(0020,000D)", "Y", "N"),
        ("accession_number", "VARCHAR", "", "RIS accession number", "(0008,0050)", "N", "N"),
        ("study_id", "VARCHAR", "", "Study ID", "(0020,0010)", "N", "N"),
        ("study_datetime", "TIMESTAMP_NTZ", "", "Combined study date/time", "", "N", "N"),
        ("study_date", "DATE", "", "Date study started", "(0008,0020)", "N", "N"),
        ("study_time", "TIME", "", "Time study started", "(0008,0030)", "N", "N"),
        ("study_description", "VARCHAR", "", "Study description", "(0008,1030)", "N", "N"),
        ("referring_physician", "VARCHAR", "", "Referring physician name", "(0008,0090)", "N", "Y"),
        ("admitting_diagnosis", "VARCHAR", "", "Admitting diagnosis description", "(0008,1080)", "N", "N"),
        ("study_instance_uid_root", "VARCHAR", "", "UID root identifier", "", "N", "N"),
        ("number_of_series", "INTEGER", "", "Count of series in study", "(0020,1206)", "N", "N"),
        ("number_of_instances", "INTEGER", "", "Count of instances in study", "(0020,1208)", "N", "N"),
        ("modalities_in_study", "ARRAY", "", "List of modalities in study", "(0008,0061)", "N", "N"),
        ("_source_file", "VARCHAR", "", "Source file path", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_series", "Core Hierarchy", "Group of instances with common acquisition context", "FK: study_key → dicom_study", [
        ("series_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("study_key", "INTEGER", "FK → dicom_study(study_key)", "Parent study", "", "Y", "N"),
        ("series_instance_uid", "VARCHAR", "NOT NULL UNIQUE", "Globally unique series identifier", "(0020,000E)", "Y", "N"),
        ("series_number", "INTEGER", "", "Series number", "(0020,0011)", "N", "N"),
        ("modality", "VARCHAR(16)", "NOT NULL", "Equipment type: CT, MR, CR, DX, US, NM, PT, XA, MG, SEG, SR, OT", "(0008,0060)", "Y", "N"),
        ("body_part_examined", "VARCHAR", "", "Body part examined", "(0018,0015)", "N", "N"),
        ("laterality", "VARCHAR(16)", "", "L, R, or B", "(0020,0060)", "N", "N"),
        ("series_description", "VARCHAR", "", "Series description", "(0008,103E)", "N", "N"),
        ("frame_of_reference_uid", "VARCHAR", "", "Spatial coordinate frame", "(0020,0052)", "N", "N"),
        ("patient_position", "VARCHAR(16)", "", "Patient positioning (HFS, FFS, etc.)", "(0018,5100)", "N", "N"),
        ("performed_station_name", "VARCHAR", "", "Station where acquired", "", "N", "N"),
        ("performed_location", "VARCHAR", "", "Location where acquired", "", "N", "N"),
        ("series_date", "DATE", "", "Date series acquired", "(0008,0021)", "N", "N"),
        ("series_time", "TIME", "", "Time series acquired", "(0008,0031)", "N", "N"),
        ("protocol_name", "VARCHAR", "", "Acquisition protocol name", "(0018,1030)", "N", "N"),
        ("number_of_instances", "INTEGER", "", "Count of instances in series", "(0020,1209)", "N", "N"),
        ("_source_file", "VARCHAR", "", "Source file path", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_instance", "Core Hierarchy", "Individual DICOM SOP instance (image, SR, SEG, RT, etc.)", "FK: series_key → dicom_series", [
        ("instance_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("series_key", "INTEGER", "FK → dicom_series(series_key)", "Parent series", "", "Y", "N"),
        ("sop_instance_uid", "VARCHAR", "NOT NULL UNIQUE", "Globally unique instance identifier", "(0008,0018)", "Y", "N"),
        ("sop_class_uid", "VARCHAR", "NOT NULL", "SOP Class identifier (type of object)", "(0008,0016)", "Y", "N"),
        ("instance_number", "INTEGER", "", "Image number in series", "(0020,0013)", "N", "N"),
        ("image_type", "ARRAY", "", "Image type components", "(0008,0008)", "N", "N"),
        ("acquisition_datetime", "TIMESTAMP_NTZ", "", "Combined acquisition date/time", "", "N", "N"),
        ("content_datetime", "TIMESTAMP_NTZ", "", "Combined content date/time", "", "N", "N"),
        ("acquisition_date", "DATE", "", "Date of acquisition", "(0008,0022)", "N", "N"),
        ("acquisition_time", "TIME", "", "Time of acquisition", "(0008,0032)", "N", "N"),
        ("content_date", "DATE", "", "Date content created", "(0008,0023)", "N", "N"),
        ("content_time", "TIME", "", "Time content created", "(0008,0033)", "N", "N"),
        ("number_of_frames", "INTEGER", "DEFAULT 1", "Number of frames (multi-frame)", "(0028,0008)", "N", "N"),
        ("specific_character_set", "VARCHAR", "", "Character encoding", "(0008,0005)", "N", "N"),
        ("burned_in_annotation", "VARCHAR", "", "Whether annotations are burned in", "(0028,0301)", "N", "Y"),
        ("presentation_intent", "VARCHAR", "", "FOR PROCESSING or FOR PRESENTATION", "(0008,0068)", "N", "N"),
        ("file_path", "VARCHAR", "", "Original file path", "", "N", "N"),
        ("file_size_bytes", "INTEGER", "", "File size in bytes", "", "N", "N"),
        ("transfer_syntax_uid", "VARCHAR", "", "DICOM transfer syntax", "(0002,0010)", "N", "N"),
        ("_source_file", "VARCHAR", "", "Source file path", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_frame", "Core Hierarchy", "Per-frame metadata for multi-frame objects (enhanced CT/MR, SEG)", "FK: instance_key → dicom_instance", [
        ("frame_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Parent instance", "", "Y", "N"),
        ("frame_number", "INTEGER", "NOT NULL", "Frame number (1-based)", "", "Y", "N"),
        ("frame_content_datetime", "TIMESTAMP_NTZ", "", "Frame acquisition time", "(0018,9074)", "N", "N"),
        ("image_position_patient", "ARRAY", "", "3D position in patient coordinates", "(0020,0032)", "N", "N"),
        ("image_orientation_patient", "ARRAY", "", "Direction cosines", "(0020,0037)", "N", "N"),
        ("slice_location", "FLOAT", "", "Slice location in mm", "(0020,1041)", "N", "N"),
        ("temporal_position_index", "INTEGER", "", "Temporal position for dynamic scans", "(0020,9128)", "N", "N"),
        ("cardiac_cycle_position", "FLOAT", "", "Position in cardiac cycle", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_equipment", "Technical Context", "Acquisition device/scanner details", "FK: series_key → dicom_series", [
        ("equipment_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("series_key", "INTEGER", "FK → dicom_series(series_key)", "Parent series", "", "Y", "N"),
        ("manufacturer", "VARCHAR", "", "Device manufacturer", "(0008,0070)", "N", "N"),
        ("manufacturer_model_name", "VARCHAR", "", "Model name", "(0008,1090)", "N", "N"),
        ("device_serial_number", "VARCHAR", "", "Serial number", "(0018,1000)", "N", "N"),
        ("software_versions", "VARCHAR", "", "Software versions", "(0018,1020)", "N", "N"),
        ("institution_name", "VARCHAR", "", "Institution name", "(0008,0080)", "N", "N"),
        ("institution_address", "VARCHAR", "", "Institution address", "(0008,0081)", "N", "N"),
        ("station_name", "VARCHAR", "", "Station identifier", "(0008,1010)", "N", "N"),
        ("institutional_dept_name", "VARCHAR", "", "Department name", "(0008,1040)", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_image_pixel", "Technical Context", "Pixel grid and encoding characteristics (not pixel values)", "FK: instance_key → dicom_instance", [
        ("image_pixel_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Parent instance", "", "Y", "N"),
        ("image_rows", "INTEGER", "", "Image height in pixels", "(0028,0010)", "N", "N"),
        ("image_columns", "INTEGER", "", "Image width in pixels", "(0028,0011)", "N", "N"),
        ("number_of_frames", "INTEGER", "", "Number of frames", "(0028,0008)", "N", "N"),
        ("samples_per_pixel", "INTEGER", "", "Color channels (1=grayscale, 3=RGB)", "(0028,0002)", "N", "N"),
        ("photometric_interpretation", "VARCHAR", "", "MONOCHROME1, MONOCHROME2, RGB, etc.", "(0028,0004)", "N", "N"),
        ("bits_allocated", "INTEGER", "", "Bits allocated per pixel", "(0028,0100)", "N", "N"),
        ("bits_stored", "INTEGER", "", "Bits stored per pixel", "(0028,0101)", "N", "N"),
        ("high_bit", "INTEGER", "", "Most significant bit", "(0028,0102)", "N", "N"),
        ("pixel_representation", "INTEGER", "", "0=unsigned, 1=signed", "(0028,0103)", "N", "N"),
        ("planar_configuration", "INTEGER", "", "Pixel data organization", "(0028,0006)", "N", "N"),
        ("rescale_intercept", "FLOAT", "", "Rescale intercept (CT HU)", "(0028,1052)", "N", "N"),
        ("rescale_slope", "FLOAT", "", "Rescale slope (CT HU)", "(0028,1053)", "N", "N"),
        ("window_center", "ARRAY", "", "Window center values", "(0028,1050)", "N", "N"),
        ("window_width", "ARRAY", "", "Window width values", "(0028,1051)", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_image_plane", "Technical Context", "Spatial resolution and positioning", "FK: instance_key → dicom_instance, frame_key → dicom_frame", [
        ("image_plane_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Parent instance", "", "Y", "N"),
        ("frame_key", "INTEGER", "FK → dicom_frame(frame_key)", "Parent frame (if multi-frame)", "", "N", "N"),
        ("pixel_spacing", "ARRAY", "", "Row/column spacing in mm", "(0028,0030)", "N", "N"),
        ("slice_thickness", "FLOAT", "", "Slice thickness in mm", "(0018,0050)", "N", "N"),
        ("image_position_patient", "ARRAY", "", "3D position in patient coordinates", "(0020,0032)", "N", "N"),
        ("image_orientation_patient", "ARRAY", "", "Direction cosines (6 values)", "(0020,0037)", "N", "N"),
        ("spacing_between_slices", "FLOAT", "", "Spacing between slices in mm", "(0018,0088)", "N", "N"),
        ("position_reference_indicator", "VARCHAR", "", "Anatomical reference point", "(0020,1040)", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_procedure_step", "Workflow Context", "Mapping to requested/performed procedures and codes", "FK: study_key → dicom_study", [
        ("procedure_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("study_key", "INTEGER", "FK → dicom_study(study_key)", "Parent study", "", "Y", "N"),
        ("requested_procedure_id", "VARCHAR", "", "Requested procedure identifier", "(0040,1001)", "N", "N"),
        ("requested_procedure_description", "VARCHAR", "", "Requested procedure description", "(0032,1060)", "N", "N"),
        ("requested_procedure_code_seq", "VARIANT", "", "Coded procedure request (JSON)", "(0032,1064)", "N", "N"),
        ("performed_procedure_step_id", "VARCHAR", "", "Performed step identifier", "(0040,0253)", "N", "N"),
        ("performed_procedure_description", "VARCHAR", "", "Performed procedure description", "(0040,0254)", "N", "N"),
        ("performed_procedure_type", "VARCHAR", "", "Type of performed procedure", "", "N", "N"),
        ("performed_procedure_code_seq", "VARIANT", "", "Coded performed procedure (JSON)", "(0040,0260)", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_dose_summary", "Dose & Exposure", "Summarized exposure parameters for CT/radiography analytics", "FK: series_key → dicom_series, study_key → dicom_study", [
        ("dose_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("series_key", "INTEGER", "FK → dicom_series(series_key)", "Parent series", "", "Y", "N"),
        ("study_key", "INTEGER", "FK → dicom_study(study_key)", "Parent study", "", "Y", "N"),
        ("ctdi_vol", "FLOAT", "", "CT Dose Index Volume (mGy)", "(0018,9345)", "N", "N"),
        ("dose_length_product", "FLOAT", "", "Dose Length Product (mGy*cm)", "(0018,9346)", "N", "N"),
        ("exposure_time", "FLOAT", "", "Exposure time (ms)", "(0018,1150)", "N", "N"),
        ("kvp", "FLOAT", "", "Peak kilovoltage (kVp)", "(0018,0060)", "N", "N"),
        ("xray_tube_current", "FLOAT", "", "X-ray tube current (mA)", "(0018,1151)", "N", "N"),
        ("exposure", "FLOAT", "", "Exposure (mAs)", "(0018,1152)", "N", "N"),
        ("acquisition_protocol", "VARCHAR", "", "Acquisition protocol name", "(0018,1030)", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_segmentation_metadata", "Derived Objects", "Metadata describing segmentation objects (SEG SOP)", "FK: instance_key → dicom_instance, referenced_series_key → dicom_series", [
        ("segmentation_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Parent SEG instance", "", "Y", "N"),
        ("referenced_series_key", "INTEGER", "FK → dicom_series(series_key)", "Referenced source series", "", "N", "N"),
        ("segment_number", "INTEGER", "", "Segment number", "(0062,0004)", "N", "N"),
        ("segment_label", "VARCHAR", "", "Segment label (e.g., Liver, Tumor)", "(0062,0005)", "N", "N"),
        ("segment_description", "VARCHAR", "", "Segment description", "(0062,0006)", "N", "N"),
        ("segmentation_type", "VARCHAR", "", "BINARY or FRACTIONAL", "(0062,0001)", "N", "N"),
        ("segmentation_fractional_type", "VARCHAR", "", "PROBABILITY or OCCUPANCY", "(0062,0010)", "N", "N"),
        ("recommended_display_grayscale", "VARIANT", "", "Display parameters (JSON)", "", "N", "N"),
        ("anatomic_region_code_seq", "VARIANT", "", "Anatomic region code (JSON)", "(0008,2218)", "N", "N"),
        ("property_category_code_seq", "VARIANT", "", "Property category (JSON)", "(0062,0003)", "N", "N"),
        ("property_type_code_seq", "VARIANT", "", "Property type (JSON)", "(0062,000F)", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_structured_report_header", "Derived Objects", "High-level metadata for Structured Reports (SR SOP)", "FK: instance_key → dicom_instance", [
        ("sr_header_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Parent SR instance", "", "Y", "N"),
        ("completion_flag", "VARCHAR", "", "COMPLETE or PARTIAL", "(0040,A491)", "N", "N"),
        ("verification_flag", "VARCHAR", "", "VERIFIED or UNVERIFIED", "(0040,A493)", "N", "N"),
        ("document_title", "VARCHAR", "", "SR document title", "(0042,0010)", "N", "N"),
        ("coding_scheme_identification", "VARIANT", "", "Coding scheme references (JSON)", "(0008,0110)", "N", "N"),
        ("referenced_instance_keys", "ARRAY", "", "Referenced SOP instances", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_file_location", "Storage", "Where each DICOM object is stored and how to access it", "FK: instance_key → dicom_instance", [
        ("location_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Parent instance", "", "Y", "N"),
        ("storage_uri", "VARCHAR", "", "Full URI to DICOM object", "", "N", "N"),
        ("storage_provider", "VARCHAR", "", "Storage provider (S3, GCS, Azure, Snowflake)", "", "N", "N"),
        ("storage_container", "VARCHAR", "", "Bucket/container name", "", "N", "N"),
        ("object_key", "VARCHAR", "", "Object key/path within container", "", "N", "N"),
        ("transfer_syntax_uid", "VARCHAR", "", "Transfer syntax of stored object", "(0002,0010)", "N", "N"),
        ("checksum", "VARCHAR", "", "File checksum (MD5/SHA256)", "", "N", "N"),
        ("ingestion_source", "VARCHAR", "", "How file was ingested (PACS, upload, etc.)", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_element", "Generic Store", "Generic store for any DICOM data element (long-tail and private tags)", "FK: instance_key → dicom_instance", [
        ("element_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Parent instance", "", "Y", "N"),
        ("frame_number", "INTEGER", "", "Frame number if per-frame element", "", "N", "N"),
        ("tag_group", "INTEGER", "", "DICOM tag group (e.g., 0010)", "", "Y", "N"),
        ("tag_element", "INTEGER", "", "DICOM tag element (e.g., 0020)", "", "Y", "N"),
        ("tag_hex", "VARCHAR(8)", "", "Hex representation (e.g., 00100020)", "", "N", "N"),
        ("name", "VARCHAR", "", "Element name from data dictionary", "", "N", "N"),
        ("vr", "VARCHAR(4)", "", "Value representation (LO, PN, DA, etc.)", "", "N", "N"),
        ("vm", "INTEGER", "", "Value multiplicity", "", "N", "N"),
        ("value_string", "VARCHAR", "", "String value of element", "", "N", "N"),
        ("value_number", "FLOAT", "", "Numeric value of element", "", "N", "N"),
        ("value_datetime", "TIMESTAMP_NTZ", "", "Date/time value of element", "", "N", "N"),
        ("value_binary_ref", "VARCHAR", "", "Reference to binary data", "", "N", "N"),
        ("is_private", "BOOLEAN", "", "Whether this is a private tag", "", "N", "N"),
        ("private_creator", "VARCHAR", "", "Private creator identifier", "", "N", "N"),
        ("sequence_item_key", "INTEGER", "", "Parent sequence item (if nested)", "", "N", "N"),
        ("sequence_path", "VARCHAR", "", "Path in sequence hierarchy", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("dicom_sequence_item", "Generic Store", "Represents a single item within a sequence (SQ) element", "FK: instance_key → dicom_instance, parent_element_key → dicom_element", [
        ("sequence_item_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Parent instance", "", "Y", "N"),
        ("parent_element_key", "INTEGER", "FK → dicom_element(element_key)", "Parent sequence element", "", "Y", "N"),
        ("item_index", "INTEGER", "", "Item index within sequence (0-based)", "", "N", "N"),
        ("sequence_path", "VARCHAR", "", "Path in sequence hierarchy", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("image_embedding", "ML/AI", "Vector representations of images for similarity search and ML", "FK: instance_key → dicom_instance, frame_key → dicom_frame, model_key → embedding_model", [
        ("embedding_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("instance_key", "INTEGER", "FK → dicom_instance(instance_key)", "Source image instance", "", "Y", "N"),
        ("frame_key", "INTEGER", "FK → dicom_frame(frame_key)", "Source frame (if multi-frame)", "", "N", "N"),
        ("segmentation_key", "INTEGER", "FK → dicom_segmentation_metadata(segmentation_key)", "Source segmentation", "", "N", "N"),
        ("embedding_vector", "ARRAY", "", "Embedding vector (float array)", "", "Y", "N"),
        ("model_key", "INTEGER", "FK → embedding_model(model_key)", "Model used to generate embedding", "", "Y", "N"),
        ("representation_scope", "VARCHAR", "", "FULL_IMAGE, ROI, PATCH, etc.", "", "N", "N"),
        ("representation_version", "VARCHAR", "", "Embedding version identifier", "", "N", "N"),
        ("created_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "When embedding was generated", "", "N", "N"),
        ("source_image_uri", "VARCHAR", "", "URI to source image", "", "N", "N"),
    ]),
    ("embedding_model", "ML/AI", "Catalog of embedding models used to generate image embeddings", "", [
        ("model_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("model_name", "VARCHAR", "NOT NULL", "Model name", "", "Y", "N"),
        ("model_version", "VARCHAR", "", "Model version", "", "N", "N"),
        ("modality_scope", "ARRAY", "", "Modalities this model handles (CT, MR, etc.)", "", "N", "N"),
        ("task_scope", "ARRAY", "", "Tasks: classification, detection, segmentation", "", "N", "N"),
        ("dimensionality", "INTEGER", "", "Embedding vector dimensionality", "", "Y", "N"),
        ("training_data_summary", "VARCHAR", "", "Description of training data", "", "N", "N"),
        ("preprocessing_spec", "VARIANT", "", "Preprocessing configuration (JSON)", "", "N", "N"),
        ("postprocessing_notes", "VARCHAR", "", "Post-processing notes", "", "N", "N"),
        ("owning_team", "VARCHAR", "", "Team responsible for model", "", "N", "N"),
        ("regulatory_notes", "VARCHAR", "", "Regulatory/compliance notes", "", "N", "N"),
        ("_loaded_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "ETL load timestamp", "", "N", "N"),
    ]),
    ("embedding_evaluation", "ML/AI", "Track quality metrics for embedding models on representative datasets", "FK: model_key → embedding_model", [
        ("evaluation_key", "INTEGER", "AUTOINCREMENT PRIMARY KEY", "Surrogate key", "", "Y", "N"),
        ("model_key", "INTEGER", "FK → embedding_model(model_key)", "Evaluated model", "", "Y", "N"),
        ("dataset_name", "VARCHAR", "", "Evaluation dataset name", "", "N", "N"),
        ("dataset_description", "VARCHAR", "", "Dataset description", "", "N", "N"),
        ("metric_name", "VARCHAR", "", "Metric name (AUC, accuracy, cosine_sim, etc.)", "", "Y", "N"),
        ("metric_value", "FLOAT", "", "Metric value", "", "Y", "N"),
        ("metric_details", "VARIANT", "", "Additional metric details (JSON)", "", "N", "N"),
        ("evaluated_at", "TIMESTAMP_NTZ", "DEFAULT CURRENT_TIMESTAMP()", "When evaluation was performed", "", "N", "N"),
    ]),
]

headers = ["column_name", "data_type", "constraints", "description", "dicom_tag", "required_for_search", "contains_phi"]

ws = wb.active
ws.title = "Data Model Overview"

ws.append(["DICOM Data Model Reference - 18 Tables"])
ws.merge_cells("A1:G1")
ws["A1"].font = Font(bold=True, size=14, color="2F5496")
ws.append([])

ws.append(["Table Name", "Category", "Description", "Relationships", "Column Count"])
for col in range(1, 6):
    cell = ws.cell(row=3, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i, (name, category, desc, rels, cols) in enumerate(tables):
    ws.append([name, category, desc, rels, len(cols)])
    for col in range(1, 6):
        ws.cell(row=4+i, column=col).border = thin_border

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 15

ws_detail = wb.create_sheet("Column Details")
detail_headers = ["table_name", "category", "table_description", "column_name", "data_type",
                  "constraints", "description", "dicom_tag", "required_for_search", "contains_phi"]
ws_detail.append(detail_headers)
for col_idx, h in enumerate(detail_headers, 1):
    cell = ws_detail.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

row = 2
for name, category, desc, rels, cols in tables:
    for col_data in cols:
        col_name, dtype, constraints, col_desc, tag, req, phi = col_data
        ws_detail.append([name, category, desc, col_name, dtype, constraints, col_desc, tag, req, phi])
        for c in range(1, 11):
            ws_detail.cell(row=row, column=c).border = thin_border
        row += 1

ws_detail.column_dimensions['A'].width = 35
ws_detail.column_dimensions['B'].width = 18
ws_detail.column_dimensions['C'].width = 55
ws_detail.column_dimensions['D'].width = 35
ws_detail.column_dimensions['E'].width = 18
ws_detail.column_dimensions['F'].width = 35
ws_detail.column_dimensions['G'].width = 55
ws_detail.column_dimensions['H'].width = 15
ws_detail.column_dimensions['I'].width = 18
ws_detail.column_dimensions['J'].width = 15

ws_search = wb.create_sheet("Search Corpus (Flat)")
search_headers = ["search_text", "table_name", "column_name", "data_type", "dicom_tag",
                  "category", "description", "constraints", "contains_phi", "relationships"]
ws_search.append(search_headers)
for col_idx, h in enumerate(search_headers, 1):
    cell = ws_search.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill

row = 2
for name, category, desc, rels, cols in tables:
    for col_data in cols:
        col_name, dtype, constraints, col_desc, tag, req, phi = col_data
        search_text = f"Table {name} ({category}): {desc}. Column {col_name} ({dtype}) - {col_desc}."
        if tag:
            search_text += f" DICOM tag {tag}."
        if phi == "Y":
            search_text += " Contains PHI - requires masking."
        if constraints:
            search_text += f" Constraints: {constraints}."
        ws_search.append([search_text, name, col_name, dtype, tag, category, col_desc, constraints, phi, rels])
        row += 1

ws_search.column_dimensions['A'].width = 100
ws_search.column_dimensions['B'].width = 35
ws_search.column_dimensions['C'].width = 35
ws_search.column_dimensions['D'].width = 18
ws_search.column_dimensions['E'].width = 15
ws_search.column_dimensions['F'].width = 18
ws_search.column_dimensions['G'].width = 55
ws_search.column_dimensions['H'].width = 35
ws_search.column_dimensions['I'].width = 15
ws_search.column_dimensions['J'].width = 50

output_path = "/Users/mgandhirajan/Documents/CoCo/HCLS/coco-healthcare-skills/references/dicom_data_model_reference.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Generated: {output_path}")
print(f"Tables: {len(tables)}")
total_cols = sum(len(cols) for _, _, _, _, cols in tables)
print(f"Total columns: {total_cols}")
print(f"Sheets: Data Model Overview, Column Details, Search Corpus (Flat)")
